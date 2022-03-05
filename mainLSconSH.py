from hdbcli import dbapi
import requests
from getpass import getuser 
import pandas as pd
import numpy as np
import urllib3
from openpyxl import load_workbook
from pathlib import Path
import datetime
import time
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import os


def entregas_sin_coordinar():
    user = getuser()
    hoy = datetime.datetime.now()
    dia = hoy.day
    
    if dia <= 9:
        dia = "0" + str(dia)
    else:
        dia  
    
    mes=hoy.month
    if mes <= 9:
        mes = "0" + str(mes)
    else:
        mes  
    
    agno = hoy.year
    hora = hoy.hour
    minutos = hoy.minute
    fh_corrida = str(agno)+ "-" +str(mes)+ "-" +str(dia) + "-" +str(hora) + "-" +str(minutos)
    
    #----------------------------- Fecha limite de propuesta de entrega en pedido
    current_date = datetime.datetime.today()
    new_date = current_date + datetime.timedelta(3)
    new_date.strftime('%Y%m%d')
    day = new_date.day
    if day <= 9:
        day = "0" + str(dia)
    else:
        day
    month = new_date.month
    if month <= 9:
        month = "0" + str(month)
    else:
        month
    year = new_date.year

    fechaLimite = f"{year}{month}{day}"
    # fechaLimite = "20220304"
    print("---")

    conn=dbapi.connect(address="172.31.0.138",
                        port="30115",
                        user="OYP",
                        password="5tAgt7S8k7XvDx",
                        sslValidateCertificate=False)
    cursor=conn.cursor()
    cursor.execute("SET SCHEMA SAPABAP1")
    
    
    # conn =dbapi.connect(
    #     address="172.31.0.130",
    #     port="30115",
    #     user="OYP",
    #     password="A112ShhtPLZYVv",
    #     sslValidateCertificate=False
    #     ) ##PRD
    cursor=conn.cursor()
    cursor.execute("SET SCHEMA SAPABAP1") 
    
    # ------------------------------------------- PEDIDOS PENDIENTES -------------------------------------------------------------#
    bloqueos = ["PE", "PR"]
    l_bloqueos = tuple(bloqueos)

    clientes_excluir = ['60000006','60000122','60000110']
    l_excluir = tuple(clientes_excluir)


    query1 = f""" 
    
    SELECT 
    VBUP.VBELN AS VBUP_VBELN,    
    VBUK.VBELN AS VBUK_VBELN,
    VBAP.POSNR,
    VBAP.MATNR,
    VBAP.VBELN AS VBAP_VBELN,
    MARA.ZZTXT, 
    MARA.ZZESTADO,
    VBAK.ERDAT,
    VBAK.VDATU,
    VBAK.KUNNR,
    VBAK.VBELN AS VBAK_VBELN,
    ZSD_CONVENIOS.DESCRIPCION,
    ZSD_CONVENIOS.CONVENIO,
    ZSD_CONVENIOS.PORC_AFIL,
    KNA1.NAME1,
    KNA1.KUNNR

    FROM   
    (VBUP
    inner join VBUK 
    on  VBUK.VBELN = VBUP.VBELN
    inner join VBAP
    on  VBAP.POSNR = VBUP.POSNR
    and VBAP.VBELN = VBUP.VBELN
    inner join MARA
    on  MARA.MATNR = VBAP.MATNR
    inner join VBAK
    on  VBAK.VBELN = VBAP.VBELN
    inner join ZSD_CONVENIOS
    on  ZSD_CONVENIOS.CONVENIO = VBAK.ZZCONVENIO
    inner join KNA1
    on  KNA1.KUNNR = VBAK.KUNNR )

    WHERE 
    (VBUP.LFSTA = 'A' OR VBUP.LFSTA = 'B') AND 
    (VBUP.ABSTA = 'A' OR VBUP.ABSTA = 'B') AND 
    (ZSD_CONVENIOS.PORC_AFIL = '0') AND
    (VBAK.AUART = 'ZCAP' OR VBAK.AUART = 'ZTER' OR VBAK.AUART = 'ZTRA') AND
    (VBAK.LIFSK IN {l_bloqueos}) AND
    (VBAK.KUNNR NOT IN {l_excluir}) AND
    (VBAK.VDATU <= {fechaLimite})
    """
    # KNA1.ZZAGENTE = '05'

    cursor.execute(query1)
    df1 = pd.read_sql_query(query1,conn)
    print("---------------Query 1: Pedidos Pendientes ---------"+"\n")
    print(f"Directorio Actual: {os.getcwd()}")
    print(df1.head(20))
    print(df1.shape)
    df1.to_excel(f"C:/Users/{getuser()}/Desktop/OyP/Coordinacion Pedidos/ExcelSQL/query1.xlsx")

# #------------------------ SEGUNDA QUERY ------------------------# 
#     l_pedidos = []
#     for i in range(0,10000000):
#         try:
#             ped = df1.loc[i,'VBUP_VBELN']
#             l_pedidos.append(ped)
#         except:
#             break

#     ids_pedidos = (', '.join("'" + item + "'" for item in l_pedidos)) ### Formato lindo para SQL parametro
#     # ------------------------------------------- AFILIADOS DE PEDIDOS PENDIENTES -------------------------------------------------------------#
#     query2 = """
#     SELECT 
#     VBPA.KUNNR,
#     VBPA.VBELN AS VBPA_VBELN
    
#     FROM VBPA
#     WHERE VBPA.PARVW = 'ZA' AND VBPA.VBELN IN (%s)
#     """ %(ids_pedidos) 
    
#     #############################################ULTIMO AGREGADO: Se modifica WE por SH poruqe no se determinan en los Pedidos
#     query2_1 = """
#     SELECT 
#     VBPA.KUNNR,
#     VBPA.VBELN AS VBPA_VBELN,
#     KNA1.ZZAGENTE
#     FROM (VBPA
#     inner join KNA1
#     on KNA1.KUNNR = VBPA.KUNNR)

#     WHERE VBPA.PARVW = 'WE' AND VBPA.VBELN IN (%s)
#     """%(ids_pedidos) 

#     # select VBPA~VBELN VBPA~PARVW KNA1~ZZAGENTE KNA1~KUNNR
#     # from ( VBPA
#     #        inner join KNA1
#     #        on  KNA1~KUNNR = VBPA~KUNNR )

#     print("--------------- Query 2: Afiliados de pedidos pendientes ---------"+"\n")     
#     cursor.execute(query2)
#     df2 = pd.read_sql_query(query2,conn)
#     print(df2.head(10))
#     print(df2.shape)
#     print("--------------- Query 2_1: Saco centros asistenciales: NO TRAE NADA CON  ---------"+"\n")
#     df2_1 = pd.read_sql_query(query2_1,conn)
#     print(df2_1.head(10))
#     print(df2_1.shape)
#     indexNames = df2_1[ df2_1['ZZAGENTE'] == '04' ].index
#     indexNames1 = df2_1[ df2_1['ZZAGENTE'] == '06' ].index
#     indexNames2 = df2_1[ df2_1['ZZAGENTE'] == '07' ].index
#     indexNames3 = df2_1[ df2_1['ZZAGENTE'] == '08' ].index
#     indexNames4 = df2_1[ df2_1['ZZAGENTE'] == '02' ].index
#     indexNames5 = df2_1[ df2_1['ZZAGENTE'] == '01' ].index
#     indexNames6 = df2_1[ df2_1['ZZAGENTE'] == '03' ].index
#     indexNames7 = df2_1[ df2_1['ZZAGENTE'] == '09' ].index
#     # Delete these row indexes from dataFrame
#     df2_1.drop(indexNames , inplace=True)
#     df2_1.drop(indexNames1 , inplace=True)
#     df2_1.drop(indexNames2 , inplace=True)
#     df2_1.drop(indexNames3 , inplace=True)
#     df2_1.drop(indexNames4 , inplace=True)
#     df2_1.drop(indexNames5 , inplace=True)
#     df2_1.drop(indexNames6 , inplace=True)
#     df2_1.drop(indexNames7 , inplace=True)
#     print(df2_1.head(10))
#     print(df2_1.shape)
#     #df2.to_excel("C:/Users/LDelgado/Desktop/oyp/df/query2.xlsx")
    
#     l_pedidos_afiliados = []
#     for i in range(0,10000000):
#         try:
#             afi = df2.loc[i,'KUNNR']
#             l_pedidos_afiliados.append(afi)
#         except:
#             break

#     #ids_pedidos_afiliados = (', '.join("'" + item + "'" for item in l_pedidos_afiliados)) ### Formato lindo para SQL parametro
#     ids_pedidos_afiliados = tuple(l_pedidos_afiliados)

#     print("---------------Merge 1: Pedidos pend + afi ---------"+"\n")
#     df12 = pd.merge(left=df1, right=df2, left_on='VBAK_VBELN', right_on='VBPA_VBELN')
#     print(df12.head(10))
#     print(df12.shape)
#     print("---------------Merge 1.2: Pedidos pend + afi (con filtro centro asistencial)---------"+"\n")
#     df13 = pd.merge(left=df2_1, right=df12, left_on='VBPA_VBELN', right_on='VBAK_VBELN')
#     print(df13.head(10))
#     print(df13.shape)
#     #df12.to_excel("C:/Users/LDelgado/Desktop/oyp/df/pedidos_y_afi.xlsx")

#     base_pedidos = df13[['KUNNR_y','VBPA_VBELN_x','POSNR','MATNR']]
#     #     indexNames = df[ df['Stock'] == 'No' ].index
#     # # Delete these row indexes from dataFrame
#     # df.drop(indexNames , inplace=True)
#     #base_pedidos.to_excel("C:/Users/LDelgado/Desktop/oyp/df/base_pedidos-"+fh_corrida+".xlsx", sheet_name="pedidos")

#     wb = load_workbook('pedidos_candidatos.xlsx')
#     ws = wb['Base']
#     for index, row in df13[['KUNNR_y']].iterrows():
#         cell = f'A{index + 16}'
#         ws[cell] = int(row[0])
#     for index, row in df13[['VBPA_VBELN_x']].iterrows():
#         cell = f'B{index + 16}'
#         ws[cell] = int(row[0])
#     for index, row in df13[['POSNR']].iterrows():
#         cell = f'C{index + 16}'
#         ws[cell] = int(row[0])
#     for index, row in df13[['MATNR']].iterrows():
#         cell = f'D{index + 16}'
#         ws[cell] = int(row[0])
#     #wb.save('pedidos_candidatos.xlsx')
    
#     # ------------------------------------------- ENTREGAS -------------------------------------------------------------#
#     query3 =f"""
#     SELECT 
#     VBUP.POSNR,
#     VBUP.WBSTA,
#     VBUP.FKSTA,
#     VBUP.KOSTA,
#     VBUP.KOQUA,
#     LIPS.VBELN AS VBELN1,
#     LIPS.POSNR AS POSNR1,
#     LIPS.PSTYV,
#     LIPS.ERZET,
#     LIPS.ERDAT,
#     LIPS.MATNR AS MATNR1,
#     LIPS.WERKS,
#     LIPS.LGORT,
#     LIPS.PRODH,
#     LIPS.LFIMG,
#     LIPS.VRKME,
#     LIPS.VGBEL,
#     LIPS.VGPOS,
#     LIPS.KCMENG,
#     LIPS.MEINS,
#     LIKP.ERNAM,
#     LIKP.ERZET,
#     LIKP.ERDAT,
#     LIKP.VSTEL,
#     LIKP.LFART,
#     LIKP.LFDAT,
#     LIKP.KODAT,
#     LIKP.ROUTE,
#     LIKP.FAKSK,
#     LIKP.LIFSK,
#     LIKP.VBTYP,
#     LIKP.LPRIO,
#     LIKP.VSBED,
#     LIKP.KUNNR AS ID_DEST,
#     LIKP.KUNAG,
#     LIKP.WADAT_IST,
#     LIKP.XBLNR,
#     LIKP.ZZESTADO,
#     LIKP.ZZTRAZABLE,
#     LIKP.ZZESTADO_INT,
#     LIKP.ZZTURNO,
#     LIKP.ZZOL_HORA,
#     LIKP.ZZOL_FECHA,
#     LIKP.ZZOL_TEXTO,
#     LIKP.ZZOL_ESTADO ,
#     LIKP.ZZOL_NRO_GUIA,
#     LIKP.ZZURGENTE,
#     LIKP.ZZOPORTUNIDAD,
#     LIKP.ZZDESTINO,
#     LIKP.ZZLETRA,
#     LIKP.ZZEST_LISTO_FACT,
#     LIKP.ZZLISTO_FACTURAR,
#     LIKP.ZZFACTURABLE,
#     LIKP.ZZESTADO_RED,
#     LIKP.ZZCOND_FRIO,
#     LIKP.ZZESTADO_ANMAT,
#     LIKP.ZZANEXO,
#     LIKP.ZZDOKA,
#     LIKP.ZZTICKET_DOKA,
#     LIKP.VBELN,
#     VBPA.PARVW,
#     VBPA.VBELN AS ENTREGA,
#     VBPA.KUNNR AS ID_AFI
#     FROM VBUP
#     inner join LIPS
#     on  LIPS.POSNR = VBUP.POSNR
#     and LIPS.VBELN = VBUP.VBELN
#     inner join LIKP
#     on  LIKP.VBELN = LIPS.VBELN
#     inner join VBPA
#     on  VBPA.VBELN = LIKP.VBELN
#     WHERE
#     ((VBUP.WBSTA = 'C') AND
#     (LIPS.PSTYV <> 'ZAFI') AND
#     (LIPS.PSTYV <> 'ZLOT') AND
#     (LIPS.PSTYV <> 'ZAF1') AND
#     (LIPS.PSTYV <> 'ZNRD') AND
#     (LIPS.PSTYV <> 'ZPOS') AND
#     (VBPA.PARVW = 'ZA') AND 
#     (LIKP.KUNNR LIKE '0084%')) AND 
#     (LIKP.KUNNR <> '0085041336') AND
#     (VBPA.KUNNR IN {ids_pedidos_afiliados}) AND ((LIKP.LFART = 'ZEVT') OR (LIKP.LFART = 'ZLIA') OR (LIKP.LFART = 'ZENC')) AND
#     (LIKP.LFDAT > '20200101')
#     """
    
#     print("---------------Query 3: Entregas por afiliado con pedidos pendientes ---------"+"\n")     
#     df3 = pd.read_sql_query(query3,conn)
#     print(df3.head(10))
#     print(df3.shape)
#     #df3.to_excel("C:/Users/LDelgado/Desktop/oyp/df/query3.xlsx")

#     base_entregas = df3[['ENTREGA','MATNR1','ID_AFI','ID_DEST','POSNR1']]
#     #base_entregas.to_excel("C:/Users/LDelgado/Desktop/oyp/df/base_entregas-"+fh_corrida+".xlsx",sheet_name="entregas")
#     cursor.execute(query3)
#     cursor.close()
#     conn.close()

#     for index, row in df3[['ENTREGA']].iterrows():
#         cell = f'O{index + 16}'
#         ws[cell] = int(row[0])
#     for index, row in df3[['MATNR1']].iterrows():
#         cell = f'P{index + 16}'
#         ws[cell] = int(row[0])
#     for index, row in df3[['ID_AFI']].iterrows():
#         cell = f'Q{index + 16}'
#         ws[cell] = int(row[0])
#     for index, row in df3[['ID_DEST']].iterrows():
#         cell = f'R{index + 16}'
#         ws[cell] = int(row[0])
#     for index, row in df3[['POSNR1']].iterrows():
#         cell = f'S{index + 16}'
#         ws[cell] = int(row[0])
#     wb.save('pedidos_candidatos_final.xlsx')

#     print("--"*30)
#     print("Fin")

if __name__=="__main__":    
    entregas_sin_coordinar()
    # WE que arranque con 84 --> fueron a fcia



